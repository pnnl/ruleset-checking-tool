import glob
import json
import os
import re

import pandas as pd
from openpyxl import utils
from openpyxl.styles import Alignment, Font, PatternFill
from rct229.utils.natural_sort import natural_keys


def create_rule_test_documentation_spreadsheet(
    ruleset_standard, test_json_branch="master"
):
    """Generates an Excel documentation file for all ruletest JSONS found for a particular ruleset standard in:
    ruleset_checking_tool/rct229/ruletest_engine/ruletest_jsons/RULESET_STANDARD.

    Resulting Excel file will be titled RULESET_STANDARD_rules.xlsx in the directory mentioned above.

    Parameters
    ----------

    ruleset_standard : str

        A string representing a ruleset standard directory as specified under:
            ruleset_checking_tool/rct229/ruletest_engine/ruletest_jsons/RULESET_STANDARD
        Example: "ashrae902019"

    test_json_branch : str

        This string points to the name of the branch that the test JSON hyperlinks will reference. This helps freeze or
        lockdown the documentation to a fixed branch rather than having it move along with master, if desired.

        Example: "RT/JG/schema_update_016_017"

    """

    ruletest_directory = f"../{ruleset_standard}"
    excel_path = f"{ruletest_directory}/{ruleset_standard}_rules.xlsx"

    # Aggregate rule test information into a dictionary
    master_ruletest_dict = generate_rule_test_dictionary(
        ruleset_standard, test_json_branch
    )

    # Style and generate Excel spreadsheet from aggregated rule test information
    write_rule_test_excel_from_dictionary(master_ruletest_dict, excel_path)

    print(f"Test JSONs written to excel at directory: '{excel_path}'")


def natural_sort_key(rule_unit_test):
    """Split the Rule_Unit_Test into numeric parts for natural ordering."""
    parts = re.split(r"[-]", rule_unit_test)  # Split by '-'
    return [
        int(parts[0]),
        int(parts[1]),
        parts[2],
    ]  # Convert numeric parts to integers for proper sorting


def generate_rule_test_dictionary(ruleset_standard, test_json_branch="master"):
    """Aggregates all ruletest JSON information found in the specified directory into a dictionary for easy data
    manipulation and processing.

    Parameters
    ----------

    ruleset_standard : str
        A string representing a ruleset standard directory as specified under:
            ruleset_checking_tool/rct229/ruletest_engine/ruletest_jsons/RULESET_STANDARD
        Example: "ashrae902019"

    test_json_branch : str

        This string points to the name of the branch that the test JSON hyperlinks will reference. This helps freeze or
        lockdown the documentation to a fixed branch rather than having it move along with master, if desired.

        Example: "RT/JG/schema_update_016_017"

    Returns
    -------
    master_rule_data_dict: dict
        Dictionary containing aggregated ruletest JSON information for all different sections
        Format: master_rule_data_dict[SECTION] = { rule_data_columns[i]: [x1, x2, x3]}
            EX: master_rule_data_dict["section23"] = {"Rule": ["23-1", "23-1", ...],
                                                      "Rule_Unit_Test": ["23-1-a", "23-1-b", ...]
                                                       etc.

    """

    ruletest_directory = f"../{ruleset_standard}"
    this_file_dir = os.path.abspath(os.path.dirname(__file__))
    ruletest_path = os.path.join(this_file_dir, ruletest_directory)
    ruletest_url = f"https://github.com/pnnl/ruleset-checking-tool/tree/{test_json_branch}/rct229/ruletest_engine/ruletest_jsons/{ruleset_standard}"

    # Dictionary mapping sections to a list of rule test JSONs relevant to that section
    # Format ruletest_dict[SECTION_NAME] = ['ruletest1.json', 'ruletest2.json', ...]
    ruletest_dict = {}

    # Use os.scandir to get directory entries
    with os.scandir(ruletest_path) as entries:
        for entry in entries:
            if entry.is_dir() and entry.name.startswith("section"):
                # Section is last subdirectory
                subdirectory = entry.path
                section = subdirectory.split("\\")[-1]

                # Initialize the dictionary as a list then populate it with rule test JSONs
                ruletest_dict[section] = []
                ruletest_dict[section].extend(
                    glob.glob(os.path.join(subdirectory, "rule*.json"))
                )

    # Reorder sections to be in numerical order (i.e., avoid section1, section11, section12, section5, section6)
    sections_list = list(ruletest_dict)
    sorted_sections = sorted(sections_list, key=lambda x: natural_keys(x))
    ruletest_dict = {key: ruletest_dict[key] for key in sorted_sections}

    # Initialize dict to hold Rules information.
    # Keys include: Rule, Rule_Unit_Test, Test_Description, Expected_Rule_Outcome, Rule_Unit_Test_JSON
    master_rule_data_dict = {}
    rule_data_columns = [
        "Rule",
        "Rule_Unit_Test",
        "Test_Description",
        "Expected_Rule_Outcome",
        "Rule_Unit_Test_JSON",
    ]

    # Aggregate JSON information into master_rule_data_dict
    # Format: master_rule_data_dict[SECTION] = { rule_data_columns[i]: [x1, x2, x3]}
    #     EX: master_rule_data_dict["section23"] = {"Rule": ["23-1", "23-1", ...],
    #                                               "Rule_Unit_Test": ["23-1-a", "23-1-b", ...]
    #                                                etc.
    for section_key in ruletest_dict:

        # Initialize each column as a list in rule_data_dict
        for rule_column in rule_data_columns:
            if section_key not in master_rule_data_dict:
                master_rule_data_dict[section_key] = {}
            master_rule_data_dict[section_key][rule_column] = []

        # Read JSON files and populate master_rule_data_dict
        for ruletest_json in ruletest_dict[section_key]:
            with open(ruletest_json, "r") as json_file:
                ruletest_json_dict = json.load(json_file)

            # Append this test JSONs case information
            for test_case in ruletest_json_dict:
                case_dict = ruletest_json_dict[test_case]
                section = case_dict["Section"]
                rule = case_dict["Rule"]
                master_rule_data_dict[section_key]["Rule"].append(f"{section}-{rule}")
                master_rule_data_dict[section_key]["Rule_Unit_Test"].append(
                    f"{section}-{rule}-{case_dict['Test']}"
                )
                master_rule_data_dict[section_key]["Test_Description"].append(
                    case_dict["test_description"]
                )
                master_rule_data_dict[section_key]["Expected_Rule_Outcome"].append(
                    case_dict["expected_rule_outcome"]
                )
                master_rule_data_dict[section_key]["Rule_Unit_Test_JSON"].append(
                    f"{ruletest_url}/section{section}/rule_{section}_{rule}.json"
                )

        # Zipping all corresponding elements together to sort by rule test name (e.g., 5-2-a should be before 5-10-a)
        zipped = zip(
            master_rule_data_dict[section_key]["Rule"],
            master_rule_data_dict[section_key]["Rule_Unit_Test"],
            master_rule_data_dict[section_key]["Test_Description"],
            master_rule_data_dict[section_key]["Expected_Rule_Outcome"],
            master_rule_data_dict[section_key]["Rule_Unit_Test_JSON"],
        )

        # Sorting the zipped items using natural numeric-alphabetical order of Rule_Unit_Test (element 1)
        sorted_data = sorted(zipped, key=lambda x: natural_sort_key(x[1]))

        # Reconstruct dictionary-of-lists from sorted tuples
        master_rule_data_dict[section_key] = {
            "Rule": [item[0] for item in sorted_data],
            "Rule_Unit_Test": [item[1] for item in sorted_data],
            "Test_Description": [item[2] for item in sorted_data],
            "Expected_Rule_Outcome": [item[3] for item in sorted_data],
            "Rule_Unit_Test_JSON": [item[4] for item in sorted_data],
        }

    return master_rule_data_dict


def write_rule_test_excel_from_dictionary(master_ruletest_dict, excel_path):

    """Takes an aggregates ruletest information dict and generates a styled Excel spreadsheet documenting all the
    information contained in said dictionary.

    Parameters
    ----------

    master_ruletest_dict : dict
        Dictionary of aggregated ruletest information. Generated by the generate_rule_test_dictionary function

    excel_path: str
        The path where you'd write out the rule test documentation Excel file.
        Ex: 'C:/rct_repo/rct229/ruletest_engine/ruletest_jsons/ashrae9012019/ashrae9012019_rules.xlsx'

    """

    # Set column widths
    column_widths = [8, 13, 70, 21, 19]

    # Set column letters for various headers
    header_to_letter = {
        "Rule": "A",
        "Rule_Unit_Test": "B",
        "Test_Description": "C",
        "Expected_Rule_Outcome": "D",
        "Rule_Unit_Test_JSON": "E",
    }

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

        # Write each section to a different sheet and style it
        for section_key in master_ruletest_dict:
            section_df = pd.DataFrame(master_ruletest_dict[section_key])
            section_df.to_excel(writer, sheet_name=section_key, index=False)

            writer.book.active = writer.sheets[section_key]
            worksheet = writer.book.active

            # Changing header font color and fill color
            for cell in worksheet[1]:
                cell.font = Font(color="000000", bold=True)
                cell.fill = PatternFill(
                    start_color="EBF1DE", end_color="EBF1DE", fill_type="solid"
                )

            # Update column widths
            for i, width in enumerate(column_widths, start=1):
                column_letter = utils.get_column_letter(i)
                worksheet.column_dimensions[column_letter].width = width

            # Set styling for table values
            for row in range(2, len(section_df) + 2):
                # Set relative cells
                rule_cell = f'{header_to_letter["Rule"]}{row}'
                rule_unit_test_cell = f'{header_to_letter["Rule_Unit_Test"]}{row}'
                test_description = f'{header_to_letter["Test_Description"]}{row}'
                rule_outcome_cell = f'{header_to_letter["Expected_Rule_Outcome"]}{row}'
                hyperlink_cell = f'{header_to_letter["Rule_Unit_Test_JSON"]}{row}'

                # Apply hyperlinks to cells in the Rule_Unit_Test_JSON column linked to URLs in the Rule column
                json_name = worksheet[f'{header_to_letter["Rule"]}{row}'].value
                display_value = f"rule{json_name}.json"
                url_value = worksheet[hyperlink_cell].value

                worksheet[
                    hyperlink_cell
                ].value = f'=HYPERLINK("{url_value}", "{display_value}")'
                worksheet[hyperlink_cell].style = "Hyperlink"
                worksheet[hyperlink_cell].font = Font(
                    color="0000FF", underline="single"
                )

                # Set cell alignments
                worksheet[rule_cell].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                worksheet[rule_unit_test_cell].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                worksheet[test_description].alignment = Alignment(wrap_text=True)
                worksheet[rule_outcome_cell].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                worksheet[hyperlink_cell].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
